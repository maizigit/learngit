# """ excel的读取库，实现测试用例的读取"""
#
# import openpyxl
#
# import time
#
# # 访问excel
# from common.base import Key
#
# wb = openpyxl.load_workbook('../Data/测试用例.xlsx')
#
# # 获取sheet页
#
# # sheet = wb['Sheet1']
# for sheets in wb.sheetnames:
#     sheet = wb[sheets]
#     # 获取单元格内容
#     for values in sheet.values:
#         # print(values)
#         #
#         if type(values[0]) is int:
#             data ={}
#             data['name'] = values[2]
#             data['value'] = values[3]
#             data['text'] = values[4]
#             # print(data)
#             # 处理参数中为None的参数
#             for key in list(data.keys()):
#                 if data[key] is None:
#                     del data[key]
#
#             print(data)
#
#             # 基于关键字和参数将操作执行起来
#             if values[1] =='open_browser':
#                 tt = Key(**data)
#
#             # elif values[1] =='open':
#             #     t.open(**data)
#             # elif values[1] =='input':
#             #     t.input(**data)
#             # elif values[1] == 'click':
#             #     t.input(**data)
#             # 优化代码用反射机制
#             else:
#                 getattr(tt,values[1])(**d

import openpyxl

class Excel:
    def __init__(self,file_name):
        self.file_name = file_name

    def get_rows_value(self,sheetname,num):
        """
        读取excel表中某一行的数据，并返回到row_list列表中
        """

        # 读取excel文件
        wb = openpyxl.load_workbook(self.file_name)

        # 获取某张sheet表
        sheet = wb[sheetname]

        # 读取表中的某行数据
        row_list = []
        for i in sheet[num]:
            row_list.append(i.value)
        return row_list


if __name__ == '__main__':
    e = Excel('../Data/白数据.xlsx')
    mydata = e.get_rows_value('税金贷',2)
    print(mydata)





